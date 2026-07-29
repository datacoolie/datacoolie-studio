from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any

import yaml

from datacoolie_studio.domains.storage.uri import StorageProviderNotEnabled, require_local_path


class MetadataReadError(RuntimeError):
    pass


def read_metadata_file(uri: str) -> dict[str, Any]:
    try:
        path = require_local_path(uri)
    except StorageProviderNotEnabled as exc:
        raise MetadataReadError(str(exc)) from exc
    if not path.exists():
        raise MetadataReadError(f"Metadata file not found: {uri}")
    try:
        return read_metadata_bytes(uri, path.read_bytes())
    except MetadataReadError:
        raise
    except Exception as exc:
        raise MetadataReadError(f"Cannot parse metadata file: {uri}") from exc


def read_metadata_bytes(uri: str, content: bytes) -> dict[str, Any]:
    suffix = Path(uri.split("?", 1)[0]).suffix.lower()
    try:
        if suffix in {".yaml", ".yml"}:
            data = yaml.safe_load(content.decode("utf-8")) or {}
        elif suffix in {".xlsx", ".xls"}:
            return _read_excel_bytes(content)
        else:
            data = json.loads(content.decode("utf-8"))
    except MetadataReadError:
        raise
    except Exception as exc:
        raise MetadataReadError(f"Cannot parse metadata file: {uri}") from exc
    if not isinstance(data, dict):
        raise MetadataReadError("Metadata root must be an object")
    return data


def _read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise MetadataReadError("JSON metadata root must be an object")
    return data


def _read_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    if not isinstance(data, dict):
        raise MetadataReadError("YAML metadata root must be an object")
    return data


def _read_excel(path: Path) -> dict[str, Any]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "connections": _sheet_rows(workbook, "connections"),
            "dataflows": _excel_dataflows(workbook),
            "schema_hints": _excel_schema_hints(workbook),
        }
    finally:
        workbook.close()


def _read_excel_bytes(content: bytes) -> dict[str, Any]:
    import openpyxl

    workbook = openpyxl.load_workbook(
        BytesIO(content), read_only=True, data_only=True
    )
    try:
        return {
            "connections": _sheet_rows(workbook, "connections"),
            "dataflows": _excel_dataflows(workbook),
            "schema_hints": _excel_schema_hints(workbook),
        }
    finally:
        workbook.close()


def _sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        return []
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        item = {headers[index]: _cast_cell(value) for index, value in enumerate(row) if index < len(headers) and headers[index]}
        rows.append({key: value for key, value in item.items() if value is not None})
    return rows


def _cast_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        if value.startswith(("{", "[")):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return value
        if "," in value:
            return [part.strip() for part in value.split(",") if part.strip()]
    return value


def _excel_dataflows(workbook: Any) -> list[dict[str, Any]]:
    dataflows = []
    for row in _sheet_rows(workbook, "dataflows"):
        df: dict[str, Any] = {}
        source: dict[str, Any] = {}
        destination: dict[str, Any] = {}
        transform: dict[str, Any] = {}
        for key, value in row.items():
            if key.startswith("source_"):
                source[key.removeprefix("source_")] = value
            elif key.startswith("destination_"):
                destination[key.removeprefix("destination_")] = value
            elif key.startswith("transform_"):
                transform[key.removeprefix("transform_")] = value
            else:
                df[key] = value
        if source:
            df["source"] = source
        if destination:
            df["destination"] = destination
        if transform:
            df["transform"] = transform
        if df:
            dataflows.append(df)
    return dataflows


def _excel_schema_hints(workbook: Any) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str | None], dict[str, Any]] = {}
    for row in _sheet_rows(workbook, "schema_hints"):
        conn = row.get("connection_name")
        table = row.get("table_name")
        schema = row.get("schema_name")
        if not conn or not table:
            continue
        key = (str(conn), str(table), str(schema) if schema else None)
        group = grouped.setdefault(key, {"connection_name": conn, "table_name": table, "hints": []})
        if schema:
            group["schema_name"] = schema
        hint = {k: v for k, v in row.items() if k not in {"connection_name", "table_name", "schema_name"}}
        if hint:
            group["hints"].append(hint)
    return list(grouped.values())
