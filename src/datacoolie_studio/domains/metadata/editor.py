from __future__ import annotations

import hashlib
import json
import shutil
import tempfile
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from sqlalchemy import select
from sqlalchemy.orm import Session

from datacoolie_studio.core.config import backup_dir
from datacoolie_studio.core.identity import name_to_uuid
from datacoolie_studio.db.models import (
    EnvironmentMetadataEditorDraft,
    EnvironmentSource,
    MetadataBackup,
    MetadataEditorDraft,
    MetadataSaveEvent,
)
from datacoolie_studio.domains.metadata.reader import MetadataReadError
from datacoolie_studio.domains.metadata.reader import read_metadata_bytes
from datacoolie_studio.domains.metadata.storage_io import (
    conditional_replace,
    conditional_create,
    create_local_backup,
    current_revision as storage_current_revision,
    read_source_bytes,
    storage_for_source,
)
from datacoolie_studio.domains.credentials.store import CredentialSecretStore
from datacoolie_studio.domains.storage.errors import StorageConflictError
from datacoolie_studio.domains.sources.storage_binding import (
    apply_binding,
    binding_from_source,
)
from datacoolie_studio.domains.storage.uri import join_uri, parse_storage_uri

CONNECTION_COLUMNS = [
    "connection_id",
    "name",
    "description",
    "connection_type",
    "format",
    "catalog",
    "database",
    "configure",
    "secrets_ref",
    "is_active",
]

DATAFLOW_COLUMNS = [
    "dataflow_id",
    "name",
    "description",
    "stage",
    "group_number",
    "execution_order",
    "processing_mode",
    "is_active",
    "configure",
    "source_connection_name",
    "source_schema_name",
    "source_table",
    "source_query",
    "source_python_function",
    "source_watermark_columns",
    "source_filter_expression",
    "source_configure",
    "transform_deduplicate_columns",
    "transform_latest_data_columns",
    "transform_filter_expression",
    "transform_additional_columns",
    "transform_schema_hints",
    "transform_select_columns",
    "transform_drop_columns",
    "transform_rename_columns",
    "transform_value_rules",
    "transform_hash_columns",
    "transform_masking_rules",
    "transform_configure",
    "destination_connection_name",
    "destination_schema_name",
    "destination_table",
    "destination_load_type",
    "destination_merge_keys",
    "destination_partition_columns",
    "destination_configure",
]

SCHEMA_HINT_COLUMNS = [
    "connection_name",
    "table_name",
    "schema_name",
    "column_name",
    "data_type",
    "format",
    "precision",
    "scale",
    "default_value",
    "ordinal_position",
    "is_active",
]

REQUIRED_SHEETS = ("connections", "dataflows", "schema_hints")
FILE_RUNTIME_FIELDS = {
    "connections": {"connection_id"},
    "dataflows": {"dataflow_id"},
    "schema_hints": set(),
}
STUDIO_ROUTING_COLUMNS = (
    "__metadata_source_id",
    "__metadata_source_name",
    "__metadata_source_uri",
    "__metadata_source_kind",
)
VISIBLE_STUDIO_ROUTING_COLUMN = "__metadata_source_name"


def load_editor_document(source: EnvironmentSource) -> dict[str, Any]:
    if source.source_kind != "metadata":
        raise MetadataReadError("Source is not a metadata source")
    path = Path(source.uri).expanduser()
    return _load_editor_document_from_path(
        path,
        source_id=source.id,
        environment_id=source.environment_id,
        project_id=source.environment.project_id if source.environment else None,
        source_uri=source.uri,
    )


def load_editor_document_from_raw(
    source: EnvironmentSource,
    raw_metadata: dict[str, Any],
    revision: dict[str, Any],
) -> dict[str, Any]:
    if source.source_kind != "metadata":
        raise MetadataReadError("Source is not a metadata source")
    path = Path(source.uri).expanduser()
    raw_sheets = _stringify_complex_values(_raw_metadata_to_sheets(raw_metadata))
    _materialize_default_active_flags(raw_sheets)
    _materialize_generated_ids(raw_sheets)
    _apply_studio_routing_fields(raw_sheets, source)
    document = {
        "source": {
            "source_id": source.id,
            "environment_id": source.environment_id,
            "project_id": source.environment.project_id if source.environment else None,
            "uri": source.uri,
            "name": _metadata_source_name(source),
            "format": _metadata_format(path),
            "revision": revision,
            "record_counts": _metadata_sheet_counts(raw_sheets),
        },
        "sheets": {
            name: {
                "columns": _ordered_columns(name, raw_sheets.get(name, [])),
                "rows": raw_sheets.get(name, []),
            }
            for name in REQUIRED_SHEETS
        },
    }
    document["issues"] = validate_editor_document(document)["issues"]
    return document


def load_backup_editor_document(session: Session, backup_id: int) -> dict[str, Any]:
    backup = _metadata_backup(session, backup_id)
    path = Path(backup.backup_path).expanduser()
    return _load_editor_document_from_path(
        path,
        source_id=backup.source_id,
        environment_id=backup.environment_id,
        project_id=backup.project_id,
        source_uri=backup.source_uri,
        backup_id=backup.id,
    )


def restore_backup(
    session: Session,
    backup_id: int,
    expected_revision: dict[str, Any],
    confirm_restore: bool,
    *,
    secret_store: CredentialSecretStore | None = None,
) -> dict[str, Any]:
    if not confirm_restore:
        raise MetadataReadError("Restoring metadata requires confirmation")

    backup = _metadata_backup(session, backup_id)
    source = session.get(EnvironmentSource, backup.source_id)
    if source is None or source.source_kind != "metadata":
        raise MetadataReadError("Metadata source not found")

    storage = storage_for_source(
        session, source, secret_store=secret_store, writable=True
    )
    current_bytes, current_revision = read_source_bytes(storage, source)
    if not _same_revision(current_revision, expected_revision):
        _record_save_event(session, source, "conflict", "Source file changed before restore", current_revision, None, None)
        raise MetadataConflictError("Source file changed before restore")

    backup_path = Path(backup.backup_path).expanduser()
    preview = _load_editor_document_from_path(
        backup_path,
        source_id=source.id,
        environment_id=source.environment_id,
        project_id=source.environment.project_id if source.environment else None,
        source_uri=source.uri,
        backup_id=backup.id,
    )
    validation = validate_editor_document(preview)
    if validation["status"] == "error":
        raise MetadataValidationError("Backup has validation errors", validation)

    current_backup_path = create_local_backup(
        source, current_bytes, current_revision
    )
    restored_bytes = backup_path.read_bytes()
    try:
        restored_revision = conditional_replace(
            storage, source, restored_bytes, current_revision
        )
    except StorageConflictError as exc:
        _record_save_event(
            session,
            source,
            "conflict",
            "Source object changed before restore",
            current_revision,
            None,
            current_backup_path,
        )
        raise MetadataConflictError("Source object changed before restore") from exc

    event = _record_save_event(
        session,
        source,
        "restored",
        f"Metadata restored from backup {backup.id}",
        current_revision,
        restored_revision,
        current_backup_path,
    )
    session.add(
        MetadataBackup(
            project_id=source.environment.project_id if source.environment else 0,
            environment_id=source.environment_id,
            source_id=source.id,
            source_uri=source.uri,
            backup_path=str(current_backup_path),
            source_revision_json=json.dumps(current_revision, sort_keys=True),
            saved_revision_json=json.dumps(restored_revision, sort_keys=True),
            save_event_id=event.id,
        )
    )
    draft = _latest_draft(session, source.id)
    if draft is not None:
        session.delete(draft)
    session.commit()
    return load_editor_document_from_raw(
        source,
        read_metadata_bytes(source.uri, restored_bytes),
        restored_revision,
    )


def _load_editor_document_from_path(
    path: Path,
    *,
    source_id: int,
    environment_id: int,
    project_id: int | None,
    source_uri: str,
    backup_id: int | None = None,
    ) -> dict[str, Any]:
    revision = source_revision(str(path))
    raw_sheets = _read_editor_sheets(path)
    _materialize_default_active_flags(raw_sheets)
    _materialize_generated_ids(raw_sheets)
    _apply_studio_routing_fields(
        raw_sheets,
        None,
        source_id=source_id,
        source_uri=source_uri,
        source_name=Path(source_uri).name or source_uri,
    )
    document = {
        "source": {
            "source_id": source_id,
            "environment_id": environment_id,
            "project_id": project_id,
            "uri": source_uri,
            "name": Path(source_uri).name or source_uri,
            "format": _metadata_format(path),
            "revision": revision,
            **({"backup_id": backup_id, "backup_path": str(path)} if backup_id is not None else {}),
        },
        "sheets": {
            name: {
                "columns": _ordered_columns(name, raw_sheets.get(name, [])),
                "rows": raw_sheets.get(name, []),
            }
            for name in REQUIRED_SHEETS
        },
    }
    document["issues"] = validate_editor_document(document)["issues"]
    return document


def validate_editor_document(document: dict[str, Any]) -> dict[str, Any]:
    sheets = document.get("sheets") or {}
    connections = _sheet_rows(sheets, "connections")
    dataflows = _sheet_rows(sheets, "dataflows")
    schema_hints = _sheet_rows(sheets, "schema_hints")
    issues: list[dict[str, Any]] = []

    connection_names: set[str] = set()
    seen_connections: dict[str, int] = {}
    for index, row in enumerate(connections):
        name = _string(row.get("name"))
        if not name:
            issues.append(_issue("error", "connections", index, "name", "Connection name is required"))
            continue
        if name in seen_connections:
            issues.append(_issue("error", "connections", index, "name", f"Duplicate connection name: {name}"))
        seen_connections[name] = index
        connection_names.add(name)
        if not _string(row.get("connection_id")):
            issues.append(_issue("warning", "connections", index, "connection_id", "Blank connection_id will be generated from name"))
        _validate_json_like_columns(issues, "connections", index, row, ["configure"])

    seen_dataflows: dict[str, int] = {}
    for index, row in enumerate(dataflows):
        name = _string(row.get("name"))
        if not name:
            issues.append(_issue("error", "dataflows", index, "name", "Dataflow name is required"))
        elif name in seen_dataflows:
            issues.append(_issue("error", "dataflows", index, "name", f"Duplicate dataflow name: {name}"))
        if name:
            seen_dataflows[name] = index
        if name and not _string(row.get("dataflow_id")):
            issues.append(_issue("warning", "dataflows", index, "dataflow_id", "Blank dataflow_id will be generated from name"))
        for field in ("source_connection_name", "destination_connection_name"):
            value = _string(row.get(field))
            if value and value not in connection_names:
                issues.append(_issue("error", "dataflows", index, field, f"Unknown connection: {value}"))
        if not any(_string(row.get(field)) for field in ("destination_table", "destination_configure", "transform")):
            issues.append(_issue("warning", "dataflows", index, "destination_table", "Dataflow has no clear destination table/configure/transform"))
        _validate_json_like_columns(
            issues,
            "dataflows",
            index,
            row,
            [
                "configure",
                "source_configure",
                "transform_additional_columns",
                "transform_schema_hints",
                "transform_select_columns",
                "transform_drop_columns",
                "transform_rename_columns",
                "transform_value_rules",
                "transform_hash_columns",
                "transform_masking_rules",
                "transform_configure",
                "destination_partition_columns",
                "destination_configure",
            ],
        )

    for index, row in enumerate(schema_hints):
        connection_name = _string(row.get("connection_name"))
        if not connection_name:
            issues.append(_issue("error", "schema_hints", index, "connection_name", "Connection name is required"))
        elif connection_name not in connection_names:
            issues.append(_issue("error", "schema_hints", index, "connection_name", f"Unknown connection: {connection_name}"))
        if not _string(row.get("table_name")):
            issues.append(_issue("error", "schema_hints", index, "table_name", "Table name is required"))
        if not _string(row.get("column_name")):
            issues.append(_issue("error", "schema_hints", index, "column_name", "Column name is required"))

    error_count = sum(1 for issue in issues if issue["severity"] == "error")
    warning_count = sum(1 for issue in issues if issue["severity"] == "warning")
    return {
        "status": "error" if error_count else ("warning" if warning_count else "ok"),
        "summary": {
            "errors": error_count,
            "warnings": warning_count,
            "issues": len(issues),
        },
        "issues": issues,
    }


def _has_new_validation_errors(
    validation: dict[str, Any],
    baseline_validation: dict[str, Any],
) -> bool:
    def signatures(result: dict[str, Any]) -> Counter[tuple[str, str, str]]:
        return Counter(
            (
                str(issue.get("sheet") or ""),
                str(issue.get("column") or ""),
                str(issue.get("message") or ""),
            )
            for issue in result.get("issues") or []
            if issue.get("severity") == "error"
        )

    return bool(signatures(validation) - signatures(baseline_validation))


def load_editor_draft(session: Session, source: EnvironmentSource) -> dict[str, Any] | None:
    draft = _latest_draft(session, source.id)
    if draft is None:
        return None
    return json.loads(draft.editor_document_json)


def save_editor_draft(
    session: Session,
    source: EnvironmentSource,
    document: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> dict[str, Any]:
    _assert_source_revision_current(
        session,
        source,
        document.get("source", {}).get("revision") or {},
        secret_store=secret_store,
    )
    _materialize_generated_ids(_document_sheets_as_rows(document))
    document["issues"] = validate_editor_document(document)["issues"]
    draft = _latest_draft(session, source.id)
    payload = json.dumps(document, ensure_ascii=False)
    revision_payload = json.dumps(document.get("source", {}).get("revision") or {}, sort_keys=True)
    if draft is None:
        draft = MetadataEditorDraft(source_id=source.id, base_revision_json=revision_payload, editor_document_json=payload)
        session.add(draft)
    else:
        draft.base_revision_json = revision_payload
        draft.editor_document_json = payload
    session.commit()
    return document


def delete_editor_draft(session: Session, source: EnvironmentSource) -> None:
    draft = _latest_draft(session, source.id)
    if draft is not None:
        session.delete(draft)
        session.commit()


def load_environment_editor_draft(session: Session, environment_id: int) -> dict[str, Any] | None:
    draft = _latest_environment_draft(session, environment_id)
    if draft is None:
        return None
    return json.loads(draft.editor_document_json)


def save_environment_editor_draft(
    session: Session,
    environment_id: int,
    document: dict[str, Any],
    expected_revision: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> dict[str, Any]:
    _materialize_generated_ids(_document_sheets_as_rows(document))
    document["issues"] = validate_editor_document(document)["issues"]
    payload = json.dumps(document, ensure_ascii=False)
    revision_payload = json.dumps(expected_revision or document.get("source", {}).get("revision") or {}, sort_keys=True)
    draft = _latest_environment_draft(session, environment_id)
    if draft is None:
        draft = EnvironmentMetadataEditorDraft(
            environment_id=environment_id,
            base_revision_json=revision_payload,
            editor_document_json=payload,
        )
        session.add(draft)
    else:
        draft.base_revision_json = revision_payload
        draft.editor_document_json = payload
    session.commit()
    return document


def delete_environment_editor_draft(session: Session, environment_id: int) -> None:
    draft = _latest_environment_draft(session, environment_id)
    if draft is not None:
        session.delete(draft)
        session.commit()


def list_backups(session: Session, source: EnvironmentSource) -> list[dict[str, Any]]:
    rows = session.scalars(
        select(MetadataBackup)
        .where(MetadataBackup.source_id == source.id)
        .order_by(MetadataBackup.created_at.desc(), MetadataBackup.id.desc())
    ).all()
    return [_backup_response(row) for row in rows]


def list_environment_backups(session: Session, environment_id: int) -> list[dict[str, Any]]:
    rows = session.scalars(
        select(MetadataBackup)
        .where(MetadataBackup.environment_id == environment_id)
        .order_by(MetadataBackup.created_at.desc(), MetadataBackup.id.desc())
    ).all()
    return [_backup_response(row) for row in rows]


def _backup_response(row: MetadataBackup) -> dict[str, Any]:
    return {
        "id": row.id,
        "project_id": row.project_id,
        "environment_id": row.environment_id,
        "source_id": row.source_id,
        "source_uri": row.source_uri,
        "backup_path": row.backup_path,
        "source_revision": json.loads(row.source_revision_json) if row.source_revision_json else None,
        "saved_revision": json.loads(row.saved_revision_json) if row.saved_revision_json else None,
        "created_at": row.created_at,
    }


def delete_backup(session: Session, backup_id: int) -> bool:
    backup = session.get(MetadataBackup, backup_id)
    if backup is None:
        return False
    path = Path(backup.backup_path).expanduser()
    if path.exists():
        path.unlink()
    session.delete(backup)
    session.commit()
    return True


def delete_backups(session: Session, source: EnvironmentSource) -> int:
    backups = list(
        session.scalars(
            select(MetadataBackup).where(MetadataBackup.source_id == source.id)
        ).all()
    )
    for backup in backups:
        path = Path(backup.backup_path).expanduser()
        if path.exists():
            path.unlink()
    for backup in backups:
        session.delete(backup)
    session.commit()
    return len(backups)


def delete_environment_backups(session: Session, environment_id: int) -> int:
    backups = list(
        session.scalars(
            select(MetadataBackup).where(MetadataBackup.environment_id == environment_id)
        ).all()
    )
    for backup in backups:
        path = Path(backup.backup_path).expanduser()
        if path.exists():
            path.unlink()
    for backup in backups:
        session.delete(backup)
    session.commit()
    return len(backups)


def _metadata_backup(session: Session, backup_id: int) -> MetadataBackup:
    backup = session.get(MetadataBackup, backup_id)
    if backup is None:
        raise MetadataReadError("Metadata backup not found")
    path = Path(backup.backup_path).expanduser()
    if not path.exists():
        raise MetadataReadError(f"Metadata backup file not found: {backup.backup_path}")
    return backup


def save_editor_document(
    session: Session,
    source: EnvironmentSource,
    document: dict[str, Any],
    expected_revision: dict[str, Any],
    confirm_overwrite: bool,
    *,
    validate_document: bool = True,
    secret_store: CredentialSecretStore | None = None,
    prepared_source: tuple[Any, bytes, dict[str, Any], Path] | None = None,
    serialized_bytes: bytes | None = None,
) -> dict[str, Any]:
    if not confirm_overwrite:
        raise MetadataReadError("Saving metadata requires overwrite confirmation")
    if prepared_source is None:
        storage = storage_for_source(
            session, source, secret_store=secret_store, writable=True
        )
        current_bytes, current_revision = read_source_bytes(storage, source)
        backup_path: Path | None = None
    else:
        storage, current_bytes, current_revision, backup_path = prepared_source
    if not _same_revision(current_revision, expected_revision):
        _record_save_event(session, source, "conflict", "Source file changed before save", current_revision, None, None)
        raise MetadataConflictError("Source file changed before save")

    _materialize_generated_ids(_document_sheets_as_rows(document))
    if validate_document:
        validation = validate_editor_document(document)
        if validation["status"] == "error":
            document["issues"] = validation["issues"]
            raise MetadataValidationError("Metadata document has validation errors", validation)

    backup_path = backup_path or create_local_backup(
        source, current_bytes, current_revision
    )
    serialized_bytes = serialized_bytes or _serialize_editor_document_bytes(
        document, source.uri
    )
    try:
        saved_revision = conditional_replace(
            storage,
            source,
            serialized_bytes,
            current_revision,
            verified_revision=current_revision,
        )
    except StorageConflictError as exc:
        _record_save_event(
            session,
            source,
            "conflict",
            "Source object changed before save",
            current_revision,
            None,
            backup_path,
        )
        raise MetadataConflictError("Source object changed before save") from exc
    event = _record_save_event(session, source, "saved", "Metadata source saved", current_revision, saved_revision, backup_path)
    backup = MetadataBackup(
        project_id=source.environment.project_id if source.environment else 0,
        environment_id=source.environment_id,
        source_id=source.id,
        source_uri=source.uri,
        backup_path=str(backup_path),
        source_revision_json=json.dumps(current_revision, sort_keys=True),
        saved_revision_json=json.dumps(saved_revision, sort_keys=True),
        save_event_id=event.id,
    )
    session.add(backup)
    draft = _latest_draft(session, source.id)
    if draft is not None:
        session.delete(draft)
    session.commit()
    return load_editor_document_from_raw(
        source,
        read_metadata_bytes(source.uri, serialized_bytes),
        saved_revision,
    )


def save_environment_editor_document(
    session: Session,
    environment_id: int,
    sources: list[EnvironmentSource],
    document: dict[str, Any],
    expected_revision: dict[str, Any],
    confirm_overwrite: bool,
    *,
    secret_store: CredentialSecretStore | None = None,
) -> dict[str, Any]:
    if not confirm_overwrite:
        raise MetadataReadError("Saving metadata requires overwrite confirmation")
    enabled_sources = [source for source in sources if source.enabled and source.source_kind == "metadata"]
    if not enabled_sources:
        raise MetadataReadError("No enabled metadata sources found")

    from datacoolie_studio.domains.metadata import service as metadata_service

    _materialize_generated_ids(_document_sheets_as_rows(document))
    source_revision_map = _environment_revision_sources(expected_revision)
    base_documents: dict[int, dict[str, Any]] = {}
    for source in enabled_sources:
        source_expected_revision = source_revision_map.get(source.id)
        if source_expected_revision is None:
            raise MetadataConflictError(
                f"Metadata source revision missing for source id: {source.id}"
            )
        try:
            base_documents[source.id] = (
                metadata_service.load_materialized_editor_document(
                    session,
                    source,
                    source_expected_revision,
                )
            )
        except MetadataReadError as exc:
            raise MetadataConflictError(str(exc)) from exc

    validation = validate_editor_document(document)
    if validation["status"] == "error":
        baseline_document = metadata_service.merge_environment_editor_documents(
            enabled_sources,
            list(base_documents.values()),
        )
        baseline_validation = validate_editor_document(baseline_document)
        if _has_new_validation_errors(validation, baseline_validation):
            document["issues"] = validation["issues"]
            raise MetadataValidationError(
                "Metadata document has validation errors",
                validation,
            )

    enabled_sources = _resolve_environment_document_sources(
        session,
        environment_id,
        enabled_sources,
        document,
        secret_store=secret_store,
    )

    source_documents: list[
        tuple[EnvironmentSource, dict[str, Any], dict[str, Any], bytes]
    ] = []
    for source in enabled_sources:
        base_document = base_documents.get(source.id)
        if base_document is None:
            base_document = _load_source_document_from_environment_revision(
                session,
                source,
                expected_revision,
                secret_store=secret_store,
            )
            base_documents[source.id] = base_document
        source_document = _environment_document_for_source(document, base_document, source)
        if _source_document_changed(base_document, source_document):
            source_documents.append(
                (
                    source,
                    source_document,
                    base_document.get("source", {}).get("revision") or {},
                    _serialize_editor_document_bytes(source_document, source.uri),
                )
            )

    if not source_documents:
        _assert_environment_revision_current(
            session,
            document,
            expected_revision,
            secret_store=secret_store,
        )

    originals: dict[int, tuple[Any, bytes, dict[str, Any], Path]] = {}
    for source, _, source_expected_revision, _ in source_documents:
        storage = storage_for_source(
            session, source, secret_store=secret_store, writable=True
        )
        original_bytes, original_revision = read_source_bytes(storage, source)
        if not _same_revision(original_revision, source_expected_revision):
            raise MetadataConflictError(
                f"Source file changed before save: {source.uri}"
            )
        originals[source.id] = (
            storage,
            original_bytes,
            original_revision,
            create_local_backup(source, original_bytes, original_revision),
        )

    saved_source_ids: list[int] = []
    saved_documents: dict[int, dict[str, Any]] = {}
    try:
        for source, source_document, source_expected_revision, serialized in source_documents:
            saved_documents[source.id] = save_editor_document(
                session,
                source,
                source_document,
                source_expected_revision,
                confirm_overwrite=True,
                validate_document=False,
                secret_store=secret_store,
                prepared_source=originals[source.id],
                serialized_bytes=serialized,
            )
            saved_source_ids.append(source.id)
    except Exception:
        compensation_errors: list[int] = []
        for source_id in reversed(saved_source_ids):
            source = next(
                item for item, _, _, _ in source_documents if item.id == source_id
            )
            storage, original_bytes, _, backup_path = originals[source_id]
            try:
                current = storage_current_revision(
                    storage, source, include_content_hash=True
                )
                restored = conditional_replace(
                    storage, source, original_bytes, current
                )
                _record_save_event(
                    session,
                    source,
                    "compensated",
                    "Metadata source restored after multi-source save failure",
                    current,
                    restored,
                    backup_path,
                )
            except Exception:
                compensation_errors.append(source_id)
                _record_save_event(
                    session,
                    source,
                    "partial_failure",
                    "Metadata compensation failed; recover from local backup",
                    None,
                    None,
                    backup_path,
                )
        if compensation_errors:
            raise MetadataReadError(
                "partial_failure: metadata save could not fully compensate; local backups are available"
            )
        raise

    delete_environment_editor_draft(session, environment_id)
    serialized_by_source_id = {
        source.id: serialized
        for source, _, _, serialized in source_documents
    }
    for source_id, saved_document in saved_documents.items():
        source = next(item for item in enabled_sources if item.id == source_id)
        metadata_service.store_metadata_materialization_from_bytes(
            session,
            source,
            serialized_by_source_id[source_id],
            saved_document.get("source", {}).get("revision") or {},
        )
    response_documents = [
        saved_documents.get(source.id) or base_documents[source.id]
        for source in enabled_sources
    ]
    return metadata_service.merge_environment_editor_documents(
        enabled_sources, response_documents
    )


class MetadataConflictError(MetadataReadError):
    pass


class MetadataValidationError(MetadataReadError):
    def __init__(self, message: str, validation: dict[str, Any]) -> None:
        super().__init__(message)
        self.validation = validation


def _materialize_generated_ids(sheets: dict[str, list[dict[str, Any]]]) -> None:
    for row in sheets.get("connections", []):
        name = _string(row.get("name"))
        if name and not _string(row.get("connection_id")):
            row["connection_id"] = name_to_uuid(name)
    for row in sheets.get("dataflows", []):
        name = _string(row.get("name"))
        if name and not _string(row.get("dataflow_id")):
            row["dataflow_id"] = name_to_uuid(name)


def _materialize_default_active_flags(sheets: dict[str, list[dict[str, Any]]]) -> None:
    for sheet_name in REQUIRED_SHEETS:
        for row in sheets.get(sheet_name, []):
            if "is_active" not in row or _is_empty(row.get("is_active")):
                row["is_active"] = True
            else:
                row["is_active"] = _parse_cell_value(row["is_active"])


def _document_sheets_as_rows(document: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    sheets = document.get("sheets") or {}
    return {name: _sheet_rows(sheets, name) for name in REQUIRED_SHEETS}


def _latest_draft(session: Session, source_id: int) -> MetadataEditorDraft | None:
    return session.scalars(
        select(MetadataEditorDraft)
        .where(MetadataEditorDraft.source_id == source_id)
        .order_by(MetadataEditorDraft.updated_at.desc(), MetadataEditorDraft.id.desc())
    ).first()


def _latest_environment_draft(session: Session, environment_id: int) -> EnvironmentMetadataEditorDraft | None:
    return session.scalars(
        select(EnvironmentMetadataEditorDraft)
        .where(EnvironmentMetadataEditorDraft.environment_id == environment_id)
        .order_by(EnvironmentMetadataEditorDraft.updated_at.desc(), EnvironmentMetadataEditorDraft.id.desc())
    ).first()


def _assert_source_revision_current(
    session: Session,
    source: EnvironmentSource,
    expected_revision: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> None:
    storage = storage_for_source(session, source, secret_store=secret_store)
    current_revision = storage_current_revision(
        storage, source, include_content_hash=True
    )
    if not _same_revision(current_revision, expected_revision):
        raise MetadataConflictError("Source file changed before draft save")


def _assert_environment_revision_current(
    session: Session,
    document: dict[str, Any],
    expected_revision: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> None:
    source_revision_map = _environment_revision_sources(expected_revision or document.get("source", {}).get("revision") or {})
    source_ids = _document_source_ids(document)
    missing = sorted(source_ids - set(source_revision_map))
    if missing:
        raise MetadataConflictError(f"Metadata source revision missing for source ids: {', '.join(str(item) for item in missing)}")
    for source_id in sorted(source_ids):
        source = session.get(EnvironmentSource, source_id)
        if source is None or source.source_kind != "metadata":
            raise MetadataReadError(f"Metadata source not found: {source_id}")
        _assert_source_revision_current(
            session,
            source,
            source_revision_map[source_id],
            secret_store=secret_store,
        )


def _same_revision(left: dict[str, Any], right: dict[str, Any]) -> bool:
    return str(left.get("content_hash")) == str(right.get("content_hash")) and int(left.get("size", -1)) == int(right.get("size", -2))


def _environment_revision_sources(revision: dict[str, Any]) -> dict[int, dict[str, Any]]:
    sources = revision.get("sources") if isinstance(revision, dict) else []
    result: dict[int, dict[str, Any]] = {}
    if not isinstance(sources, list):
        return result
    for item in sources:
        if not isinstance(item, dict):
            continue
        try:
            source_id = int(item.get("source_id"))
        except (TypeError, ValueError):
            continue
        source_revision_value = item.get("revision")
        result[source_id] = source_revision_value if isinstance(source_revision_value, dict) else item
    return result


def _document_source_ids(document: dict[str, Any]) -> set[int]:
    source_ids: set[int] = set()
    for rows in _document_sheets_as_rows(document).values():
        for row in rows:
            source_id = _row_source_id(row)
            if source_id is not None:
                source_ids.add(source_id)
    return source_ids


def _resolve_environment_document_sources(
    session: Session,
    environment_id: int,
    sources: list[EnvironmentSource],
    document: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> list[EnvironmentSource]:
    registry = {source.id: source for source in sources}
    tokens = _metadata_source_tokens(sources)
    for rows in _document_sheets_as_rows(document).values():
        for row in rows:
            source_id = _row_source_id(row)
            if source_id in registry:
                _apply_row_routing(row, registry[source_id])
                continue
            source_name = _string(row.get("__metadata_source_name"))
            if not source_name:
                raise MetadataReadError("metadata_source is required for every metadata row")
            source = tokens.get(_source_token(source_name))
            if source is None:
                source = _create_metadata_source_for_routing(
                    session,
                    environment_id,
                    source_name,
                    list(registry.values()),
                    secret_store=secret_store,
                )
                registry[source.id] = source
                for token, value in _metadata_source_tokens([source]).items():
                    tokens[token] = value
            _apply_row_routing(row, source)
    return list(registry.values())


def _metadata_source_tokens(sources: list[EnvironmentSource]) -> dict[str, EnvironmentSource]:
    result: dict[str, EnvironmentSource] = {}
    for source in sources:
        path = Path(source.uri).expanduser()
        values = [
            source.uri,
            path.name,
            path.stem,
            _metadata_source_name(source),
            source.label or "",
        ]
        for value in values:
            token = _source_token(value)
            if token and token not in result:
                result[token] = source
    return result


def _source_token(value: str) -> str:
    return value.strip().replace("\\", "/").lower()


def _apply_row_routing(row: dict[str, Any], source: EnvironmentSource) -> None:
    row["__metadata_source_id"] = source.id
    row["__metadata_source_name"] = _metadata_source_name(source)
    row["__metadata_source_uri"] = source.uri
    row["__metadata_source_kind"] = "metadata"


def _create_metadata_source_for_routing(
    session: Session,
    environment_id: int,
    source_name: str,
    existing_sources: list[EnvironmentSource],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> EnvironmentSource:
    cloud_parent = next(
        (
            source
            for source in existing_sources
            if source.storage_provider != "local"
        ),
        None,
    )
    if cloud_parent is not None:
        root_uri = _metadata_root_uri_text(cloud_parent)
        if not root_uri:
            raise MetadataReadError(
                "metadata_root_uri is required to create a cloud metadata object"
            )
        relative_name = _safe_metadata_object_name(source_name)
        uri = join_uri(root_uri, relative_name)
        source = EnvironmentSource(
            environment_id=environment_id,
            source_kind="metadata",
            uri=uri,
            label=relative_name,
            enabled=True,
            source_config_json=json.dumps(
                {
                    "discovery_mode": "metadata_path",
                    "metadata_root_uri": root_uri,
                },
                sort_keys=True,
            ),
        )
        apply_binding(source, binding_from_source(cloud_parent))
        session.add(source)
        session.flush()
        storage = storage_for_source(
            session, source, secret_store=secret_store, writable=True
        )
        empty = json.dumps(
            {"connections": [], "dataflows": [], "schema_hints": []},
            indent=2,
        ).encode("utf-8")
        try:
            conditional_create(storage, source, empty)
            session.commit()
        except Exception:
            session.rollback()
            raise
        session.refresh(source)
        return source

    base_dir = _metadata_source_base_dir(existing_sources)
    uri = _metadata_source_uri_for_new_name(source_name, base_dir)
    path = Path(uri).expanduser()
    if path.exists() and not path.is_file():
        raise MetadataReadError(f"Metadata source must be a file: {uri}")
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(
            json.dumps({"connections": [], "dataflows": [], "schema_hints": []}, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    source = EnvironmentSource(
        environment_id=environment_id,
        source_kind="metadata",
        uri=str(path),
        label=_new_metadata_source_label(source_name, path, base_dir),
        enabled=True,
        source_config_json=json.dumps(
            {
                "discovery_mode": "metadata_path",
                "metadata_root_uri": str(base_dir),
            },
            sort_keys=True,
        ),
    )
    session.add(source)
    session.commit()
    session.refresh(source)
    return source


def _safe_metadata_object_name(source_name: str) -> str:
    normalized = source_name.strip().replace("\\", "/")
    parsed = parse_storage_uri(normalized)
    if (
        parsed.provider != "local"
        or normalized.startswith("/")
        or any(part in {"", ".", ".."} for part in normalized.split("/"))
    ):
        raise MetadataReadError(
            "New metadata source name must be a relative path under metadata_root_uri"
        )
    candidate = Path(normalized)
    if not candidate.suffix:
        normalized = f"{normalized}.json"
    return normalized


def _metadata_root_uri_text(source: EnvironmentSource) -> str | None:
    try:
        config = json.loads(source.source_config_json or "{}")
    except json.JSONDecodeError:
        return None
    value = config.get("metadata_root_uri") if isinstance(config, dict) else None
    return value.strip() if isinstance(value, str) and value.strip() else None


def _metadata_source_uri_for_new_name(source_name: str, base_dir: Path) -> str:
    candidate = Path(source_name).expanduser()
    if not candidate.suffix:
        candidate = candidate.with_suffix(".json")
    if candidate.is_absolute():
        return str(candidate)
    return str(base_dir / candidate)


def _metadata_source_base_dir(existing_sources: list[EnvironmentSource]) -> Path:
    for source in existing_sources:
        metadata_root = _metadata_root_uri(source)
        if metadata_root:
            return metadata_root
    for source in existing_sources:
        path = Path(source.uri).expanduser()
        if path.exists() and path.is_file():
            return path.parent
        if path.suffix:
            return path.parent
    return Path.cwd()


def _metadata_root_uri(source: EnvironmentSource) -> Path | None:
    try:
        config = json.loads(source.source_config_json or "{}")
    except json.JSONDecodeError:
        return None
    if not isinstance(config, dict):
        return None
    value = config.get("metadata_root_uri")
    if not isinstance(value, str) or not value.strip():
        return None
    return Path(value).expanduser()


def _new_metadata_source_label(source_name: str, path: Path, base_dir: Path) -> str:
    raw = source_name.strip()
    if raw:
        candidate = Path(raw).expanduser()
        if not candidate.is_absolute():
            return candidate.as_posix()
    try:
        return path.relative_to(base_dir).as_posix()
    except ValueError:
        return path.name


def _metadata_sheet_counts(sheets: dict[str, list[dict[str, Any]]]) -> dict[str, int]:
    return {name: len(sheets.get(name, [])) for name in REQUIRED_SHEETS}


def _row_source_id(row: dict[str, Any]) -> int | None:
    value = row.get("__metadata_source_id")
    try:
        source_id = int(value)
    except (TypeError, ValueError):
        return None
    return source_id if source_id > 0 else None


def _load_source_document_from_environment_revision(
    session: Session,
    source: EnvironmentSource,
    expected_revision: dict[str, Any],
    *,
    secret_store: CredentialSecretStore | None = None,
) -> dict[str, Any]:
    source_revision_map = _environment_revision_sources(expected_revision)
    expected = source_revision_map.get(source.id)
    storage = storage_for_source(session, source, secret_store=secret_store)
    content, current = read_source_bytes(storage, source)
    if expected is None:
        return load_editor_document_from_raw(
            source, read_metadata_bytes(source.uri, content), current
        )
    if not _same_revision(current, expected):
        _record_save_event(session, source, "conflict", "Source file changed before environment save", current, None, None)
        raise MetadataConflictError(f"Source file changed before save: {source.uri}")
    return load_editor_document_from_raw(
        source, read_metadata_bytes(source.uri, content), current
    )


def _environment_document_for_source(
    document: dict[str, Any],
    base_document: dict[str, Any],
    source: EnvironmentSource,
) -> dict[str, Any]:
    source_document = {
        "source": dict(base_document.get("source") or {
            "source_id": source.id,
            "environment_id": source.environment_id,
            "project_id": source.environment.project_id if source.environment else None,
            "uri": source.uri,
            "name": _metadata_source_name(source),
            "format": _metadata_format(Path(source.uri).expanduser()),
            "revision": source_revision(source.uri),
        }),
        "sheets": {},
        "issues": [],
    }
    source_document["source"]["scope"] = "source"
    source_document["source"].pop("read_only", None)
    for sheet_name in REQUIRED_SHEETS:
        env_sheet = (document.get("sheets") or {}).get(sheet_name) or {}
        base_sheet = (base_document.get("sheets") or {}).get(sheet_name) or {}
        rows = [
            dict(row)
            for row in _sheet_rows(document.get("sheets") or {}, sheet_name)
            if _row_source_id(row) == source.id
        ]
        source_document["sheets"][sheet_name] = {
            "columns": _merge_source_columns(
                sheet_name,
                base_sheet.get("columns") or [],
                env_sheet.get("columns") or [],
                rows,
            ),
            "rows": rows,
        }
    source_document["issues"] = validate_editor_document(source_document)["issues"]
    return source_document


def _merge_source_columns(
    sheet_name: str,
    base_columns: list[dict[str, Any]],
    env_columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    columns: list[dict[str, Any]] = []
    seen: set[str] = set()
    row_keys = {key for row in rows for key in row}
    for column in [*base_columns, *env_columns]:
        if not isinstance(column, dict):
            continue
        key = column.get("key")
        if not key or str(key) in seen:
            continue
        if str(key).startswith("__") or key in row_keys or _is_preferred_column(str(key)):
            seen.add(str(key))
            columns.append(dict(column))
    for key in sorted(row_keys - seen):
        seen.add(key)
        columns.append({"key": key, "name": _column_display_name(key)})
    if VISIBLE_STUDIO_ROUTING_COLUMN not in seen:
        columns.append({"key": VISIBLE_STUDIO_ROUTING_COLUMN, "name": _column_display_name(VISIBLE_STUDIO_ROUTING_COLUMN)})
    return _canonical_editor_columns(sheet_name, columns, rows)


def _is_preferred_column(key: str) -> bool:
    return key in CONNECTION_COLUMNS or key in DATAFLOW_COLUMNS or key in SCHEMA_HINT_COLUMNS


def _source_document_changed(left: dict[str, Any], right: dict[str, Any]) -> bool:
    return _normalized_document_sheets(left) != _normalized_document_sheets(right)


def _normalized_document_sheets(document: dict[str, Any]) -> dict[str, Any]:
    sheets = document.get("sheets") or {}
    return {
        sheet_name: {
            "columns": [
                column.get("key")
                for column in (sheets.get(sheet_name) or {}).get("columns", [])
                if isinstance(column, dict) and column.get("key") and not str(column.get("key")).startswith("__")
            ],
            "rows": [
                {key: value for key, value in row.items() if not key.startswith("__")}
                for row in _sheet_rows(sheets, sheet_name)
            ],
        }
        for sheet_name in REQUIRED_SHEETS
    }


def _create_backup_file(source: EnvironmentSource, revision: dict[str, Any]) -> Path:
    project_id = source.environment.project_id if source.environment else 0
    target_dir = backup_dir() / f"project-{project_id}" / f"env-{source.environment_id}" / f"source-{source.id}"
    target_dir.mkdir(parents=True, exist_ok=True)
    path = Path(source.uri).expanduser()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    suffix = "".join(path.suffixes) or ".metadata"
    target = target_dir / f"{timestamp}-{revision.get('content_hash', 'unknown')[:12]}{suffix}"
    shutil.copy2(path, target)
    return target


def _record_save_event(
    session: Session,
    source: EnvironmentSource,
    status: str,
    message: str,
    source_revision_value: dict[str, Any] | None,
    saved_revision_value: dict[str, Any] | None,
    backup_path: Path | None,
) -> MetadataSaveEvent:
    event = MetadataSaveEvent(
        source_id=source.id,
        status=status,
        message=message,
        source_revision_json=json.dumps(source_revision_value, sort_keys=True) if source_revision_value else None,
        saved_revision_json=json.dumps(saved_revision_value, sort_keys=True) if saved_revision_value else None,
        backup_path=str(backup_path) if backup_path else None,
    )
    session.add(event)
    session.commit()
    return event


def _serialize_editor_document(document: dict[str, Any], path: Path) -> str:
    raw = _editor_document_to_raw_metadata(document)
    suffix = path.suffix.lower()
    if suffix in {".yaml", ".yml"}:
        try:
            from ruamel.yaml import YAML
            from io import StringIO

            yaml_writer = YAML()
            yaml_writer.default_flow_style = False
            buffer = StringIO()
            yaml_writer.dump(raw, buffer)
            return buffer.getvalue()
        except ImportError:
            return yaml.safe_dump(raw, sort_keys=False, allow_unicode=True)
    return json.dumps(raw, indent=2, ensure_ascii=False)


def _serialize_editor_document_bytes(
    document: dict[str, Any], uri: str
) -> bytes:
    suffix = Path(uri).suffix.lower()
    if suffix not in {".xlsx", ".xls"}:
        return _serialize_editor_document(document, Path(uri)).encode("utf-8")
    with tempfile.TemporaryDirectory(prefix="datacoolie-metadata-") as directory:
        target = Path(directory) / f"metadata{suffix}"
        _write_xlsx_document(document, target)
        return target.read_bytes()


def _editor_document_to_raw_metadata(document: dict[str, Any]) -> dict[str, Any]:
    sheets = document.get("sheets") or {}
    connection_columns = _sheet_column_keys(sheets, "connections")
    dataflow_columns = _sheet_column_keys(sheets, "dataflows")
    schema_hint_columns = _sheet_column_keys(sheets, "schema_hints")
    return {
        "connections": [
            _clean_row(
                _order_row(row, connection_columns),
                FILE_RUNTIME_FIELDS["connections"],
            )
            for row in _sheet_rows(sheets, "connections")
        ],
        "dataflows": [
            _unflatten_dataflow(_order_row(row, dataflow_columns))
            for row in _sheet_rows(sheets, "dataflows")
        ],
        "schema_hints": _group_schema_hints(
            [_order_row(row, schema_hint_columns) for row in _sheet_rows(sheets, "schema_hints")]
        ),
    }


def _sheet_column_keys(sheets: dict[str, Any], sheet_name: str) -> list[str]:
    sheet = sheets.get(sheet_name) or {}
    columns = sheet.get("columns") if isinstance(sheet, dict) else []
    return [
        str(column.get("key"))
        for column in columns
        if isinstance(column, dict) and column.get("key")
    ]


def _order_row(row: dict[str, Any], column_keys: list[str]) -> dict[str, Any]:
    ordered = {key: row[key] for key in column_keys if key in row}
    ordered.update({key: value for key, value in row.items() if key not in ordered})
    return ordered


def _clean_row(row: dict[str, Any], excluded_fields: set[str] | None = None) -> dict[str, Any]:
    excluded_fields = excluded_fields or set()
    result: dict[str, Any] = {}
    for key, value in row.items():
        if key.startswith("__") or key in excluded_fields or _is_empty(value) or _is_default_omitted_value(key, value):
            continue
        result[key] = _parse_cell_value(value)
    return result


def _unflatten_dataflow(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    nested: dict[str, dict[str, Any]] = {"source": {}, "transform": {}, "destination": {}}
    for key, value in row.items():
        if key.startswith("__") or key in FILE_RUNTIME_FIELDS["dataflows"] or _is_empty(value) or _is_default_omitted_value(key, value):
            continue
        target = None
        field = key
        for prefix in nested:
            marker = f"{prefix}_"
            if key.startswith(marker):
                target = prefix
                field = key[len(marker):]
                break
        parsed = _parse_cell_value(value)
        if target:
            nested[target][field] = parsed
        else:
            result[key] = parsed
    for key, value in nested.items():
        if value:
            result[key] = value
    return result


def _group_schema_hints(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    ungrouped: list[dict[str, Any]] = []
    for row in rows:
        clean = _clean_row(row)
        if not clean:
            continue
        connection_name = _string(clean.get("connection_name"))
        table_name = _string(clean.get("table_name"))
        schema_name = _string(clean.get("schema_name"))
        if not connection_name or not table_name:
            ungrouped.append(clean)
            continue
        key = (connection_name, table_name, schema_name)
        parent = grouped.setdefault(
            key,
            {
                "connection_name": connection_name,
                "table_name": table_name,
                **({"schema_name": schema_name} if schema_name else {}),
                "hints": [],
            },
        )
        hint = {
            field: value
            for field, value in clean.items()
            if field not in {"connection_name", "table_name", "schema_name"}
        }
        parent["hints"].append(hint)
    return [*grouped.values(), *ungrouped]


def _write_xlsx_document(document: dict[str, Any], path: Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    sheets = document.get("sheets") or {}
    for sheet_name in REQUIRED_SHEETS:
        sheet_data = sheets.get(sheet_name) or {}
        excluded_fields = FILE_RUNTIME_FIELDS[sheet_name]
        rows = _sheet_rows(sheets, sheet_name)
        columns = [
            column.get("key")
            for column in sheet_data.get("columns", [])
            if (
                column.get("key")
                and not str(column.get("key")).startswith("__")
                and column.get("key") not in excluded_fields
                and _xlsx_column_has_saved_values(rows, str(column.get("key")))
            )
        ]
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(columns)
        for row in rows:
            worksheet.append([_xlsx_cell_value(row.get(column), str(column)) for column in columns])
    workbook.save(path)


def _xlsx_cell_value(value: Any, key: str | None = None) -> Any:
    if key is not None and _is_default_omitted_value(key, value):
        return None
    parsed = _parse_cell_value(value)
    if isinstance(parsed, (dict, list)):
        return json.dumps(parsed, ensure_ascii=False)
    return parsed


def _parse_cell_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return None
    if stripped.lower() == "true":
        return True
    if stripped.lower() == "false":
        return False
    if stripped.startswith(("{", "[")):
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            return value
    return value


def _is_empty(value: Any) -> bool:
    return value is None or value == ""


def _is_default_omitted_value(key: str, value: Any) -> bool:
    return key == "is_active" and _parse_cell_value(value) is True


def _xlsx_column_has_saved_values(rows: list[dict[str, Any]], key: str) -> bool:
    return any(not _is_empty(row.get(key)) and not _is_default_omitted_value(key, row.get(key)) for row in rows)


def source_revision(uri: str) -> dict[str, Any]:
    path = Path(uri).expanduser()
    if not path.exists():
        raise MetadataReadError(f"Metadata file not found: {uri}")
    stat = path.stat()
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    return {
        "content_hash": digest,
        "mtime_ns": stat.st_mtime_ns,
        "size": stat.st_size,
    }


def _read_editor_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    if not path.exists():
        raise MetadataReadError(f"Metadata file not found: {path}")
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        return _stringify_complex_values(_read_xlsx_sheets(path))
    if suffix in {".yaml", ".yml"}:
        raw = _read_yaml(path)
    else:
        raw = _read_json(path)
    return _stringify_complex_values(_raw_metadata_to_sheets(raw))


def _read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise MetadataReadError("JSON metadata root must be an object")
    return data


def _read_yaml(path: Path) -> dict[str, Any]:
    try:
        from ruamel.yaml import YAML

        data = YAML(typ="safe").load(path.read_text(encoding="utf-8")) or {}
    except ImportError:
        data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(data, dict):
        raise MetadataReadError("YAML metadata root must be an object")
    return data


def _read_xlsx_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            "connections": _xlsx_sheet_rows(workbook, "connections"),
            "dataflows": _xlsx_sheet_rows(workbook, "dataflows"),
            "schema_hints": _xlsx_sheet_rows(workbook, "schema_hints"),
        }
    finally:
        workbook.close()


def _xlsx_sheet_rows(workbook: Any, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        return []
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_row]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        item = {headers[index]: _cast_cell(value) for index, value in enumerate(row) if index < len(headers) and headers[index]}
        rows.append({key: value for key, value in item.items() if value is not None})
    return rows


def _raw_metadata_to_sheets(raw: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    return {
        "connections": [dict(row) for row in raw.get("connections", []) if isinstance(row, dict)],
        "dataflows": [_flatten_dataflow(row) for row in raw.get("dataflows", []) if isinstance(row, dict)],
        "schema_hints": _flatten_schema_hints(raw.get("schema_hints", [])),
    }


def _flatten_dataflow(row: dict[str, Any]) -> dict[str, Any]:
    flattened: dict[str, Any] = {}
    for key, value in row.items():
        if key == "source" and isinstance(value, dict):
            flattened.update({f"source_{field}": field_value for field, field_value in value.items()})
        elif key == "destination" and isinstance(value, dict):
            flattened.update({f"destination_{field}": field_value for field, field_value in value.items()})
        elif key == "transform" and isinstance(value, dict):
            flattened.update({f"transform_{field}": field_value for field, field_value in value.items()})
        else:
            flattened[key] = value
    return flattened


def _stringify_complex_values(sheets: dict[str, list[dict[str, Any]]]) -> dict[str, list[dict[str, Any]]]:
    return {
        sheet_name: [
            {key: _stringify_cell_value(value) for key, value in row.items()}
            for row in rows
        ]
        for sheet_name, rows in sheets.items()
    }


def _stringify_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _flatten_schema_hints(raw_hints: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not isinstance(raw_hints, list):
        return rows
    for item in raw_hints:
        if not isinstance(item, dict):
            continue
        hints = item.get("hints")
        if isinstance(hints, list):
            base = {key: value for key, value in item.items() if key != "hints"}
            for hint in hints:
                if isinstance(hint, dict):
                    rows.append({**base, **hint})
        else:
            rows.append(dict(item))
    return rows


def _ordered_columns(sheet_name: str, rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    preferred = {
        "connections": CONNECTION_COLUMNS,
        "dataflows": DATAFLOW_COLUMNS,
        "schema_hints": SCHEMA_HINT_COLUMNS,
    }[sheet_name]
    seen = {key for row in rows for key in row}
    if sheet_name == "dataflows":
        keys = _ordered_dataflow_columns(preferred, seen)
    else:
        keys = list(preferred)
        keys.extend(sorted(key for key in seen - set(keys) if not key.startswith("__")))
    if VISIBLE_STUDIO_ROUTING_COLUMN not in keys:
        keys.append(VISIBLE_STUDIO_ROUTING_COLUMN)
    return [{"key": key, "name": _column_display_name(key)} for key in keys]


def _canonical_editor_columns(
    sheet_name: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    existing = {
        str(column["key"]): dict(column)
        for column in columns
        if isinstance(column, dict) and column.get("key")
    }
    key_sentinel = {key: None for key in existing}
    ordered = _ordered_columns(sheet_name, [*rows, key_sentinel])
    return [
        {**column, **existing.get(column["key"], {}), "key": column["key"]}
        for column in ordered
    ]


def _ordered_dataflow_columns(preferred: list[str], seen: set[str]) -> list[str]:
    keys: list[str] = []
    known = set(preferred)
    source_extra = sorted(key for key in seen - known if key.startswith("source_"))
    transform_extra = sorted(key for key in seen - known if key.startswith("transform_"))
    destination_extra = sorted(key for key in seen - known if key.startswith("destination_"))
    for key in preferred:
        keys.append(key)
        if key == "source_configure":
            keys.extend(source_extra)
        elif key == "transform_configure":
            keys.extend(transform_extra)
        elif key == "destination_configure":
            keys.extend(destination_extra)
    keys.extend(sorted(key for key in seen - set(keys) if not key.startswith("__")))
    if VISIBLE_STUDIO_ROUTING_COLUMN in seen and VISIBLE_STUDIO_ROUTING_COLUMN not in keys:
        keys.append(VISIBLE_STUDIO_ROUTING_COLUMN)
    return keys


def _column_display_name(key: str) -> str:
    if key == VISIBLE_STUDIO_ROUTING_COLUMN:
        return "metadata_source"
    return key


def _apply_studio_routing_fields(
    sheets: dict[str, list[dict[str, Any]]],
    source: EnvironmentSource | None,
    *,
    source_id: int | None = None,
    source_uri: str | None = None,
    source_name: str | None = None,
) -> None:
    if source is not None:
        source_id = source.id
        source_uri = source.uri
        source_name = _metadata_source_name(source)
    routing = {
        "__metadata_source_id": source_id,
        "__metadata_source_name": source_name or source_uri or "",
        "__metadata_source_uri": source_uri or "",
        "__metadata_source_kind": "metadata",
    }
    for rows in sheets.values():
        for row in rows:
            for key, value in routing.items():
                row[key] = value


def _metadata_source_name(source: EnvironmentSource) -> str:
    path = Path(source.uri).expanduser()
    if source.label:
        if source.label == path.stem:
            if path.parent.name.lower() in {"connection", "connections", "dataflow", "dataflows", "schema_hint", "schema_hints"}:
                return f"{path.parent.name}/{path.name}"
            metadata_root = _metadata_root_uri(source)
            if metadata_root is not None:
                try:
                    relative = path.relative_to(metadata_root).as_posix()
                except ValueError:
                    relative = ""
                if "/" in relative:
                    return relative
        return source.label
    return path.name or source.uri


def _sheet_rows(sheets: dict[str, Any], sheet_name: str) -> list[dict[str, Any]]:
    sheet = sheets.get(sheet_name) or {}
    rows = sheet.get("rows") if isinstance(sheet, dict) else []
    return [row for row in rows if isinstance(row, dict)] if isinstance(rows, list) else []


def _validate_json_like_columns(issues: list[dict[str, Any]], sheet: str, row_index: int, row: dict[str, Any], columns: list[str]) -> None:
    for column in columns:
        value = row.get(column)
        if value is None or isinstance(value, (dict, list)):
            continue
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                continue
            try:
                json.loads(stripped)
            except json.JSONDecodeError:
                issues.append(_issue("error", sheet, row_index, column, f"{column} must be valid JSON/object text"))


def _issue(severity: str, sheet: str, row_index: int, column: str, message: str) -> dict[str, Any]:
    return {
        "severity": severity,
        "sheet": sheet,
        "row_index": row_index,
        "column": column,
        "message": message,
    }


def _metadata_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".yaml", ".yml"}:
        return "yaml"
    if suffix in {".xlsx", ".xls"}:
        return "xlsx"
    return "json"


def _cast_cell(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        if value.startswith(("{", "[")):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return value
        if "," in value:
            return [part.strip() for part in value.split(",") if part.strip()]
    return value


def _string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
